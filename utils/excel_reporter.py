import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import datetime

def generate_excel_report(results, output_path="巡检报告.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "设备巡检总表"

    ws.views.sheetView[0].showGridLines = True

    headers = ["设备IP", "CPU利用率", "内存利用率", "巡检状态", "异常详情"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    alert_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    normal_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10)

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, data in enumerate(results, 2):
        row_data = [data['ip'], data['cpu'], data['memory'], data['status'], data['issue']]
        ws.append(row_data)

        current_fill = alert_fill if data['status'] == "告警" else normal_fill

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_idx == 4:
                cell.fill = current_fill

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 💡 优雅防崩兜底
    try:
        wb.save(output_path)
        print(f"📄 自动化Excel巡检报告已生成: {output_path}")
    except PermissionError:
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        alternative_path = output_path.replace(".xlsx", f"_{timestamp}.xlsx")
        wb.save(alternative_path)
        print(f"⚠️ 原报告文件被Excel占用，已自动另存为新文件: {alternative_path}")